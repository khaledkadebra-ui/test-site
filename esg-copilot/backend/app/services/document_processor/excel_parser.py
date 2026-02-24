"""
Excel/CSV Parser — ESG Copilot
================================
Reads structured Excel or CSV files and maps columns to EnergyDataInput fields.

Supports:
- Raw utility data exports (one row per meter reading)
- ESG data templates (our own pre-filled template)
- Generic spreadsheets (header-based column detection)
"""

import logging
from dataclasses import dataclass, field
from typing import Optional

logger = logging.getLogger(__name__)


@dataclass
class ExcelExtractionResult:
    rows_found: int
    sheets_processed: list[str]
    extracted_values: dict          # field_name -> aggregated value
    extraction_warnings: list[str] = field(default_factory=list)
    confidence: str = "medium"     # "high" | "medium" | "low"


# Column name aliases — map common header variants to our field names
_COLUMN_MAP: dict[str, str] = {
    # Electricity
    "electricity_kwh": "electricity_kwh",
    "electricity (kwh)": "electricity_kwh",
    "el (kwh)": "electricity_kwh",
    "elforbrug (kwh)": "electricity_kwh",
    "power consumption kwh": "electricity_kwh",
    "kwh": "electricity_kwh",
    # Natural gas
    "natural_gas_m3": "natural_gas_m3",
    "natural gas (m3)": "natural_gas_m3",
    "naturgas (m3)": "natural_gas_m3",
    "gas m3": "natural_gas_m3",
    # Diesel
    "diesel_liters": "diesel_liters",
    "diesel (liters)": "diesel_liters",
    "diesel (l)": "diesel_liters",
    "diesel litre": "diesel_liters",
    # Petrol
    "petrol_liters": "petrol_liters",
    "petrol (liters)": "petrol_liters",
    "benzin (liter)": "petrol_liters",
    # Heating oil
    "heating_oil_liters": "heating_oil_liters",
    "heating oil (liters)": "heating_oil_liters",
    "fyringsolie (liter)": "heating_oil_liters",
    # District heating
    "district_heating_kwh": "district_heating_kwh",
    "district heating (kwh)": "district_heating_kwh",
    "fjernvarme (kwh)": "district_heating_kwh",
    # Company fleet km
    "company_car_km": "company_car_km",
    "company cars (km)": "company_car_km",
    "firmabiler (km)": "company_car_km",
    "company_van_km": "company_van_km",
    "company vans (km)": "company_van_km",
    # Air travel
    "air_short_haul_km": "air_short_haul_km",
    "short haul flights (km)": "air_short_haul_km",
    "air_long_haul_km": "air_long_haul_km",
    "long haul flights (km)": "air_long_haul_km",
    # Rail
    "rail_km": "rail_km",
    "rail travel (km)": "rail_km",
    # Renewable %
    "renewable_electricity_pct": "renewable_electricity_pct",
    "renewable electricity (%)": "renewable_electricity_pct",
    "vedvarende energi (%)": "renewable_electricity_pct",
    # Employees
    "employee_count": "employee_count",
    "employees": "employee_count",
    "antal medarbejdere": "employee_count",
    "headcount": "employee_count",
    # Commuting
    "avg_commute_km_one_way": "avg_commute_km_one_way",
    "average commute km": "avg_commute_km_one_way",
    # Procurement
    "purchased_goods_spend_eur": "purchased_goods_spend_eur",
    "procurement spend (eur)": "purchased_goods_spend_eur",
    "indkøb (eur)": "purchased_goods_spend_eur",
}


class ExcelParser:
    """
    Reads XLSX, XLS, or CSV files and extracts ESG input values.

    Tries three strategies in order:
    1. ESG template detection (our pre-defined template structure)
    2. Header-based column mapping (flexible for any structured sheet)
    3. Single-value extraction (for simple one-number-per-sheet exports)
    """

    def parse(self, content: bytes, filename: str = "") -> ExcelExtractionResult:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "xlsx"
        try:
            if ext == "csv":
                return self._parse_csv(content)
            else:
                return self._parse_xlsx(content)
        except Exception as e:
            return ExcelExtractionResult(
                rows_found=0,
                sheets_processed=[],
                extracted_values={},
                extraction_warnings=[f"Parse error: {e}"],
                confidence="low",
            )

    def _parse_xlsx(self, content: bytes) -> ExcelExtractionResult:
        import openpyxl
        import io

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets_processed = []
        all_values: dict[str, float] = {}
        warnings = []
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            sheets_processed.append(sheet_name)

            # Find header row (first non-empty row)
            header_row_idx = next(
                (i for i, row in enumerate(rows) if any(cell for cell in row)),
                None,
            )
            if header_row_idx is None:
                continue

            headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]
            col_map = {}  # col_index -> our_field_name
            for i, header in enumerate(headers):
                if header in _COLUMN_MAP:
                    col_map[i] = _COLUMN_MAP[header]

            if not col_map:
                # Try single-value extraction for simple sheets (label | value layout)
                self._extract_label_value_pairs(rows, all_values, warnings)
                continue

            # Sum values across data rows
            for row in rows[header_row_idx + 1:]:
                if not any(row):
                    continue
                total_rows += 1
                for col_idx, field_name in col_map.items():
                    if col_idx < len(row) and row[col_idx] is not None:
                        try:
                            val = float(str(row[col_idx]).replace(",", "."))
                            all_values[field_name] = all_values.get(field_name, 0) + val
                        except (ValueError, TypeError):
                            pass

        confidence = "high" if len(all_values) >= 3 else "medium" if all_values else "low"
        return ExcelExtractionResult(
            rows_found=total_rows,
            sheets_processed=sheets_processed,
            extracted_values=all_values,
            extraction_warnings=warnings,
            confidence=confidence,
        )

    def _parse_csv(self, content: bytes) -> ExcelExtractionResult:
        import csv
        import io

        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return ExcelExtractionResult(0, ["csv"], {}, ["Empty CSV file"], "low")

        headers = [h.strip().lower() for h in rows[0]]
        col_map = {i: _COLUMN_MAP[h] for i, h in enumerate(headers) if h in _COLUMN_MAP}

        values: dict[str, float] = {}
        warnings = []
        row_count = 0

        for row in rows[1:]:
            if not any(cell.strip() for cell in row):
                continue
            row_count += 1
            for col_idx, field_name in col_map.items():
                if col_idx < len(row) and row[col_idx].strip():
                    try:
                        val = float(row[col_idx].replace(",", "."))
                        values[field_name] = values.get(field_name, 0) + val
                    except ValueError:
                        pass

        if not col_map:
            warnings.append("No recognised column headers found. Check the ESG data template.")

        return ExcelExtractionResult(
            rows_found=row_count,
            sheets_processed=["csv"],
            extracted_values=values,
            extraction_warnings=warnings,
            confidence="high" if len(values) >= 3 else "medium" if values else "low",
        )

    def _extract_label_value_pairs(self, rows: list, out: dict, warnings: list) -> None:
        """For sheets laid out as: [label] [value] — e.g. simple summary sheets."""
        for row in rows:
            cells = [c for c in row if c is not None]
            if len(cells) < 2:
                continue
            label = str(cells[0]).strip().lower()
            for col_label, field_name in _COLUMN_MAP.items():
                if col_label in label:
                    try:
                        val = float(str(cells[1]).replace(",", "."))
                        out[field_name] = out.get(field_name, 0) + val
                    except (ValueError, TypeError):
                        pass
                    break
